import streamlit as st
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.styles import PatternFill
from collections import defaultdict
from io import BytesIO
from tempfile import NamedTemporaryFile
import openai

st.title("🔍 Excel Cell Classification: Input, Calculation, Output")

uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

if uploaded_file:
    wb = load_workbook(filename=BytesIO(uploaded_file.read()), data_only=False)
    ws = wb.active

    openai.api_key = st.secrets.get("OPENAI_API_KEY")

    cell_types = {}
    dependencies = defaultdict(set)

    def cell_ref(cell):
        return f"{cell.column_letter}{cell.row}"

    # First pass: detect hardcoded values and formulas
    for row in ws.iter_rows():
        for cell in row:
            ref = cell_ref(cell)
            if cell.value is None:
                continue
            if cell.data_type == 'f':
                formula = str(cell.value)
                if '[' in formula and ']' in formula:
                    cell_types[ref] = 'Input (external link)'
                else:
                    tk = Tokenizer(formula)
                    for t in tk.items:
                        if t.subtype in ['range', 'operand']:
                            dependencies[ref].add(t.value)
            else:
                cell_types[ref] = 'Input (hardcoded)'

    reverse_refs = defaultdict(set)
    for target, sources in dependencies.items():
        for source in sources:
            reverse_refs[source].add(target)

    # Improved logic to detect Calculations and Outputs
    for target, sources in dependencies.items():
        for source in sources:
            reverse_refs[source].add(target)

    for row in ws.iter_rows():
        for cell in row:
            ref = f"{cell.column_letter}{cell.row}"
            if ref in cell_types:
                continue
            if cell.data_type == 'f':
                if ref in reverse_refs:
                    cell_types[ref] = 'Calculation'
                else:
                    cell_types[ref] = 'Output'
            elif cell.value is not None:
                cell_types[ref] = 'Other'

    st.write("### 🧾 Cell Classification Result")
    results = sorted(cell_types.items())
    for ref, ctype in results:
        st.write(f"**{ref}**: {ctype}")

    # Color-code the original sheet
    color_map = {
        'Input (hardcoded)': 'FFFF00',  # Yellow
        'Input (external link)': 'FFFF00',
        'Calculation': 'ADD8E6',        # Light Blue
        'Output': '90EE90',             # Light Green
        'Other': 'FFC0CB'               # Light Red
    }

    for row in ws.iter_rows():
        for cell in row:
            ref = f"{cell.column_letter}{cell.row}"
            ctype = cell_types.get(ref, 'Other')
            color = color_map.get(ctype, 'FFC0CB')
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Create Documentation Sheet
    doc_ws = wb.create_sheet(title="Documentation")
    doc_ws.append(["Overview"])
    doc_ws.append(["This spreadsheet has been automatically analyzed to classify cells as inputs, calculations, and outputs."])
    doc_ws.append([""])
    doc_ws.append(["Tab Summary"])
    doc_ws.append(["Sheet Name", "Summary"])

    for sheet in wb.sheetnames:
        if sheet in ["Documentation", "User Guide"]:
            continue
        try:
            sample_data = ""
            ws_temp = wb[sheet]
            for row in ws_temp.iter_rows(min_row=1, max_row=6, values_only=True):
                sample_data += str(row) + "\n"

            prompt = f"""
You are reviewing a sheet from an Excel workbook. Based on the following sample rows of data from the sheet titled '{sheet}', summarize what the sheet is likely doing. Consider whether it contains inputs, outputs, calculations, or reference data. Keep your summary short and clear.

Sample data:
{sample_data}
"""

            response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=[{"role": "user", "content": prompt}]
            )
            summary = response.choices[0].message.content.strip()

        except Exception as e:
            summary = f"Error generating summary: {e}"

        doc_ws.append([sheet, summary])

    # Create User Guide Sheet
    guide_ws = wb.create_sheet(title="User Guide")
    guide_ws.append(["User Guide"])
    guide_ws.append(["This sheet lists cells identified as inputs that may need to be updated manually."])
    guide_ws.append([""])
    guide_ws.append(["Cell Reference", "Input Type", "Sheet"])

    for sheet in wb.sheetnames:
        if sheet in ["Documentation", "User Guide"]:
            continue
        ws_temp = wb[sheet]
        for row in ws_temp.iter_rows():
            for cell in row:
                ref = f"{cell.column_letter}{cell.row}"
                ctype = cell_types.get(ref)
                if ctype in ["Input (hardcoded)", "Input (external link)"]:
                    guide_ws.append([ref, ctype, sheet])

    # Save and offer file for download
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        st.download_button(
            label="📥 Download Color-Coded Spreadsheet",
            data=tmp.read(),
            file_name="classified_spreadsheet.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

